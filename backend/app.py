from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import cv2
import numpy as np
from deepface import DeepFace
from collections import Counter
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from werkzeug.utils import secure_filename
import json

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
EXCEL_FILE = r'D:\Adgorithm\backend\ad_stats_log.xlsx'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload directories exist
os.makedirs(os.path.join(UPLOAD_FOLDER, 'ads'), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'qr'), exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Load OpenCV's built-in face detector
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Function to categorize age into age groups
def get_age_group(age):
    if age <= 12:
        return "Child (0-12)"
    elif age <= 19:
        return "Teenager (13-19)"
    elif age <= 34:
        return "Young Adult (20-34)"
    elif age <= 49:
        return "Middle-aged Adult (35-49)"
    else:
        return "Senior (50+)"

# Function to log detection results into an Excel file
def log_detection_to_excel(age_group, gender):
    try:
        file_name = r'D:\Adgorithm\backend\ad_stats_log.xlsx'
        print(f"Attempting to save Excel file at: {os.path.abspath(file_name)}")

        if not os.path.exists(file_name):
            # Create a new workbook and add headers
            print("Creating new Excel file...")
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "Age Group", "Gender"])
            wb.save(file_name)
            print("New Excel file created successfully")

        # Load existing workbook
        print("Loading existing Excel file...")
        wb = load_workbook(filename=file_name)
        ws = wb.active

        # Append a new row with current detection info
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([timestamp, age_group, gender])
        
        # Save the workbook
        print("Saving updated Excel file...")
        wb.save(file_name)
        print("Excel file saved successfully")
        return True
    except Exception as e:
        print(f"Error in log_detection_to_excel: {str(e)}")
        raise e  # Re-raise the exception to handle it in the calling function

# Health check route
@app.route("/", methods=["GET"])
def home():
    return jsonify({"message": "Server is running!"}), 200

# Route for predicting age and gender
@app.route("/predict-age-gender", methods=["POST"])
def predict_age_gender():
    try:
        if "image" not in request.files:
            print("No image file in request")
            return jsonify({"error": "No image file found"}), 400

        file = request.files["image"]
        if not file.filename:
            print("Empty filename")
            return jsonify({"error": "Empty filename"}), 400

        # Read the image file
        file_bytes = file.read()
        if not file_bytes:
            print("Empty file content")
            return jsonify({"error": "Empty file content"}), 400

        img = cv2.imdecode(np.frombuffer(file_bytes, np.uint8), cv2.IMREAD_COLOR)
        if img is None:
            print("Failed to decode image")
            return jsonify({"error": "Failed to decode image"}), 400

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

        if len(faces) == 0:
            print("No faces detected in the image")
            return jsonify({"error": "No faces detected"}), 400

        print(f"Detected {len(faces)} faces")

        age_groups = []
        genders = []

        for (x, y, w, h) in faces:
            face = img[y:y+h, x:x+w]

            try:
                result = DeepFace.analyze(face, actions=["age", "gender"], enforce_detection=False)[0]
                age = result.get("age", 0)
                gender_data = result.get("gender", "")
                gender = max(gender_data, key=gender_data.get) if isinstance(gender_data, dict) else str(gender_data)

                age_group = get_age_group(age)
                print(f"Face detected - Age: {age}, Gender: {gender}")
                print(f"Age group: {age_group}")

                age_groups.append(age_group)
                genders.append(gender)
            except Exception as e:
                print(f"Skipping one face due to error: {str(e)}")
                continue

        if not age_groups or not genders:
            return jsonify({"error": "Analysis failed for all detected faces"}), 500

        # Get majority values
        majority_age_group = Counter(age_groups).most_common(1)[0][0]
        majority_gender = Counter(genders).most_common(1)[0][0]

        # Log to Excel and handle any errors
        try:
            log_detection_to_excel(majority_age_group, majority_gender)
        except Exception as excel_error:
            print(f"Error logging to Excel: {str(excel_error)}")
            # Return the prediction results but include a warning about logging failure
            return jsonify({
                "age_group": majority_age_group,
                "gender": majority_gender,
                "warning": "Results not logged to Excel file"
            }), 200

        # If everything succeeded, return the normal response
        return jsonify({
            "age_group": majority_age_group,
            "gender": majority_gender
        })
    except Exception as e:
        print(f"Error in predict_age_gender: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Admin Routes
@app.route('/api/admin/stats', methods=['GET'])
def get_stats():
    try:
        # Read Excel file
        df = pd.read_excel(EXCEL_FILE)
        
        # Calculate statistics based on available columns
        stats = {
            'totalViews': len(df),
            'ageDistribution': df['Age Group'].value_counts().to_dict(),
            'genderDistribution': df['Gender'].value_counts().to_dict(),
        }
        
        # Add timeSeriesData for charts
        time_series = (
            df.groupby('date')
              .agg({'views': 'sum', 'engagement': 'sum'})
              .reset_index()
        )
        stats['timeSeriesData'] = [
            {
                'date': str(row['date']),
                'views': int(row['views']),
                'engagement': int(row['engagement'])
            }
            for _, row in time_series.iterrows()
        ]
        
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/ads', methods=['GET', 'POST'])
def manage_ads():
    if request.method == 'GET':
        try:
            # Read ads data from JSON file
            with open('uploads/ads.json', 'r') as f:
                ads = json.load(f)
            return jsonify(ads)
        except FileNotFoundError:
            return jsonify([])
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    elif request.method == 'POST':
        try:
            if 'adImage' not in request.files or 'qrImage' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400
            
            ad_image = request.files['adImage']
            qr_image = request.files['qrImage']
            
            if ad_image.filename == '' or qr_image.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            if not (allowed_file(ad_image.filename) and allowed_file(qr_image.filename)):
                return jsonify({'error': 'Invalid file type'}), 400
            
            # Generate unique filenames
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            ad_filename = f"ad_{timestamp}_{secure_filename(ad_image.filename)}"
            qr_filename = f"qr_{timestamp}_{secure_filename(qr_image.filename)}"
            
            # Save files
            ad_image.save(os.path.join(app.config['UPLOAD_FOLDER'], 'ads', ad_filename))
            qr_image.save(os.path.join(app.config['UPLOAD_FOLDER'], 'qr', qr_filename))
            
            # Save ad data
            ad_data = {
                'id': timestamp,
                'imageUrl': f"/uploads/ads/{ad_filename}",
                'qrUrl': f"/uploads/qr/{qr_filename}",
                'ageGroup': request.form.get('ageGroup'),
                'gender': request.form.get('gender'),
                'websiteLink': request.form.get('websiteLink'),
                'isActive': request.form.get('isActive', 'true').lower() == 'true'
            }
            
            try:
                with open('uploads/ads.json', 'r') as f:
                    ads = json.load(f)
            except FileNotFoundError:
                ads = []
            
            ads.append(ad_data)
            
            with open('uploads/ads.json', 'w') as f:
                json.dump(ads, f)
            
            return jsonify(ad_data), 201
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/api/admin/ads/<ad_id>', methods=['PATCH'])
def update_ad(ad_id):
    try:
        with open('uploads/ads.json', 'r') as f:
            ads = json.load(f)
        
        for ad in ads:
            if ad['id'] == ad_id:
                ad['isActive'] = request.json.get('isActive', ad['isActive'])
                break
        
        with open('uploads/ads.json', 'w') as f:
            json.dump(ads, f)
        
        return jsonify({'message': 'Ad updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/uploads/ads/<filename>')
def uploaded_ad(filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'ads'), filename)

@app.route('/uploads/qr/<filename>')
def uploaded_qr(filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'qr'), filename)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    print(f"Starting server on port {port}")
    app.run(host="0.0.0.0", port=port)
